import openpyxl
import re

def estimate_price(item_name, material, form, unit, spec):
    # Normalize inputs
    text = f"{str(item_name)} {str(material)} {str(spec)}".lower()
    form = str(form).strip()
    
    # Default price
    price = 0
    
    # Pricing Logic (Estimates in RMB)
    
    # 1. Structures / Decor
    if "桁架喷绘" in text: price = 45 # Per sqm
    elif "木质导视" in text: price = 600
    elif "咖啡车" in text: price = 3500
    elif "车贴" in text: price = 300
    elif "花艺装饰" in text: price = 1200
    elif "懒人沙发" in text: price = 80
    elif "沙滩椅" in text: price = 60
    elif "云朵气模" in text: price = 800
    elif "市集摊位" in text: price = 1000
    elif "吧台" in text: price = 2000
    elif "器皿+花艺" in text: price = 3000
    elif "高端移动厕所" in text: price = 3500
    
    # 2. Catering
    elif "鸡尾酒" in text: price = 60 # Per portion
    elif "调酒师" in text: price = 1500
    elif "主题茶歇" in text: price = 120 # Per portion
    elif "服务生" in text and "茶歇" in text: price = 400
    elif "冰淇淋" in text and "设备" in text: price = 2500
    elif "冰淇凌材料" in text: price = 35
    elif "特调咖啡" in text and "材料" in text: price = 40
    elif "工作人员" in text and ("冰淇淋" in text or "咖啡" in text): price = 400
    elif "保洁员" in text: price = 300
    
    # 3. AV / Stage
    elif "led" in text and "p3" in text: price = 300 # Per sqm
    elif "led支撑架" in text: price = 1500
    elif "屏幕高清处理器" in text: price = 1200
    elif "分屏" in text: price = 500
    elif "led控台" in text: price = 1800
    elif "大屏幕技术人员" in text: price = 1200
    elif "异形舞台" in text: price = 450 # Per sqm (Buyout)
    elif "t台" in text: price = 180 # Per sqm
    elif "地毯" in text: price = 20 # Per sqm
    elif "立体字" in text: price = 1500
    elif "鲜花置景" in text: price = 1200
    elif "演讲台" in text: price = 1000
    elif "启动仪式" in text: price = 2500
    elif "彩烟" in text: price = 1500
    elif "彩虹机" in text: price = 400
    elif "冷烟花" in text: price = 200
    elif "软包凳" in text: price = 50
    
    # Audio
    elif "线阵音响" in text: price = 800 # Per unit
    elif "补听" in text: price = 400
    elif "超低" in text: price = 600
    elif "功放" in text: price = 300
    elif "监听" in text: price = 400
    elif "线材" in text and "手麦" in text: price = 1000
    elif "线材柜" in text: price = 500
    elif "线材" in text: price = 800
    elif "音乐播放电脑" in text: price = 300
    elif "音控台" in text and "人员" not in text: price = 1500
    elif "音控台人员" in text: price = 1000
    
    # Light
    elif "光束灯" in text: price = 400
    elif "切割灯" in text: price = 600
    elif "摇头灯" in text: price = 300
    elif "爆闪灯" in text: price = 200
    elif "观众灯" in text: price = 200
    elif "电源线" in text: price = 1500
    elif "直通柜" in text: price = 800
    elif "灯光架" in text: price = 2000
    elif "信号放大器" in text: price = 300
    elif "灯光控台" in text: price = 2000
    elif "灯光师" in text: price = 1200
    
    # 4. Performers / Other
    elif "提琴" in text: price = 1500
    elif "手绘师" in text: price = 2500
    elif "乐队" in text: price = 6000
    elif "主持人" in text: price = 3500
    elif "儿童演员" in text: price = 600
    elif "高空球" in text: price = 4000
    elif "吊车" in text: price = 5000
    elif "高空杂技" in text: price = 2500
    elif "走秀模特" in text: price = 1800
    elif "化妆师" in text: price = 1200
    elif "节目背景" in text: price = 1500
    elif "摄影师" in text: price = 1500
    elif "图片直播" in text: price = 2500
    elif "摄像" in text and "单反" in text: price = 1800
    elif "航拍" in text: price = 2000
    elif "高清机位" in text: price = 2000
    elif "剪辑花絮" in text: price = 1500
    elif "礼服" in text: price = 5000
    elif "非遗簪花" in text: price = 8000
    elif "非遗鱼灯" in text: price = 150
    elif "运输费" in text: price = 1000
    elif "人工费" in text: price = 500

    return price

file_path = r'g:\learning\agent-skills\xmls\（标段1：活动）营销中心样板房开放报价清单.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

print(f"Original images count: {len(sheet._images)}")

# Columns (1-based)
# C (3): Item Name
# D (4): Material
# E (5): Spec
# F (6): Qty
# G (7): Unit
# H (8): Form
# I (9): Unit Price (Target)
# J (10): Total Price (Target)

start_row = 4 # Based on previous output, data starts row 4 (index 3 was row 4)
# Actually, let's verify row start. The header was row 2 (index 1). So data starts row 3 (index 2).
# Wait, pandas said header=1 (row 2). Data starts row 3.
# Let's iterate from row 3.

for row in range(3, sheet.max_row + 1):
    # Check if it's a data row (has Qty)
    qty_cell = sheet.cell(row=row, column=6)
    qty = qty_cell.value
    
    if qty is None or not isinstance(qty, (int, float)):
        continue
        
    item_name = sheet.cell(row=row, column=3).value
    material = sheet.cell(row=row, column=4).value
    spec = sheet.cell(row=row, column=5).value
    unit = sheet.cell(row=row, column=7).value
    form = sheet.cell(row=row, column=8).value
    
    unit_price = estimate_price(item_name, material, form, unit, spec)
    
    # Write Unit Price
    sheet.cell(row=row, column=9).value = unit_price
    
    # Write Total Price (Formula)
    # J = F * I
    # total_price = unit_price * float(qty)
    sheet.cell(row=row, column=10).value = f"=F{row}*I{row}"

new_file_path = file_path.replace('.xlsx', '_updated.xlsx')
wb.save(new_file_path)
print(f"Processing complete. Saved to {new_file_path}")

# Verify images in saved file
wb_new = openpyxl.load_workbook(new_file_path)
print(f"Saved file images count: {len(wb_new.active._images)}")
