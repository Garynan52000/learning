import openpyxl
import re

def estimate_price(item_name, material, form, unit, spec):
    # Normalize inputs
    text = f"{str(item_name)} {str(material)} {str(spec)}".lower()
    form = str(form).strip()
    
    # Base Price (Defaults to likely Rental for reusable items, Purchase for consumables)
    price = 0
    item_type = "other" # rental_based, purchase_based, labor

    # 1. Structures / Decor (Mostly Purchase/Make)
    if "桁架喷绘" in text: 
        price = 45 # Per sqm (Make)
        item_type = "purchase_based"
    elif "木质导视" in text: 
        price = 600
        item_type = "purchase_based"
    elif "咖啡车" in text: 
        price = 3500 # Rental
        item_type = "rental_based"
    elif "车贴" in text: 
        price = 300
        item_type = "purchase_based"
    elif "花艺装饰" in text: 
        price = 1200
        item_type = "purchase_based"
    elif "懒人沙发" in text: 
        price = 80 # Rental
        item_type = "rental_based"
    elif "沙滩椅" in text: 
        price = 60 # Rental
        item_type = "rental_based"
    elif "云朵气模" in text: 
        price = 800 # Rental
        item_type = "rental_based"
    elif "市集摊位" in text: 
        price = 1000 # Rental/Setup
        item_type = "rental_based"
    elif "吧台" in text: 
        price = 2000 # Rental
        item_type = "rental_based"
    elif "器皿+花艺" in text: 
        price = 3000
        item_type = "purchase_based"
    elif "高端移动厕所" in text: 
        price = 3500 # Rental
        item_type = "rental_based"
    
    # 2. Catering (Consumable/Service)
    elif "鸡尾酒" in text: price = 60; item_type = "purchase_based"
    elif "调酒师" in text: price = 1500; item_type = "labor"
    elif "主题茶歇" in text: price = 120; item_type = "purchase_based"
    elif "服务生" in text: price = 400; item_type = "labor"
    elif "冰淇淋" in text and "设备" in text: price = 2500; item_type = "rental_based"
    elif "冰淇凌材料" in text: price = 35; item_type = "purchase_based"
    elif "特调咖啡" in text and "材料" in text: price = 40; item_type = "purchase_based"
    elif "工作人员" in text: price = 400; item_type = "labor"
    elif "保洁员" in text: price = 300; item_type = "labor"
    
    # 3. AV / Stage (Mostly Rental)
    elif "led" in text and "p3" in text: price = 300; item_type = "rental_based" # Per sqm
    elif "led支撑架" in text: price = 1500; item_type = "rental_based"
    elif "屏幕高清处理器" in text: price = 1200; item_type = "rental_based"
    elif "分屏" in text: price = 500; item_type = "rental_based"
    elif "led控台" in text: price = 1800; item_type = "rental_based"
    elif "大屏幕技术人员" in text: price = 1200; item_type = "labor"
    elif "异形舞台" in text: price = 450; item_type = "purchase_based" # Make
    elif "t台" in text: price = 180; item_type = "purchase_based" # Make/Rent mixed
    elif "地毯" in text: price = 20; item_type = "purchase_based" # Consume
    elif "立体字" in text: price = 1500; item_type = "purchase_based"
    elif "鲜花置景" in text: price = 1200; item_type = "purchase_based"
    elif "演讲台" in text: price = 1000; item_type = "rental_based"
    elif "启动仪式" in text: price = 2500; item_type = "rental_based"
    elif "彩烟" in text: price = 1500; item_type = "purchase_based"
    elif "彩虹机" in text: price = 400; item_type = "rental_based"
    elif "冷烟花" in text: price = 200; item_type = "purchase_based"
    elif "软包凳" in text: price = 50; item_type = "rental_based"
    
    # Audio (Rental)
    elif "线阵音响" in text: price = 800; item_type = "rental_based"
    elif "补听" in text: price = 400; item_type = "rental_based"
    elif "超低" in text: price = 600; item_type = "rental_based"
    elif "功放" in text: price = 300; item_type = "rental_based"
    elif "监听" in text: price = 400; item_type = "rental_based"
    elif "线材" in text and "手麦" in text: price = 1000; item_type = "rental_based"
    elif "线材柜" in text: price = 500; item_type = "rental_based"
    elif "线材" in text: price = 800; item_type = "rental_based"
    elif "音乐播放电脑" in text: price = 300; item_type = "rental_based"
    elif "音控台" in text: price = 1500; item_type = "rental_based"
    
    # Light (Rental)
    elif "光束灯" in text: price = 400; item_type = "rental_based"
    elif "切割灯" in text: price = 600; item_type = "rental_based"
    elif "摇头灯" in text: price = 300; item_type = "rental_based"
    elif "爆闪灯" in text: price = 200; item_type = "rental_based"
    elif "观众灯" in text: price = 200; item_type = "rental_based"
    elif "电源线" in text: price = 1500; item_type = "rental_based"
    elif "直通柜" in text: price = 800; item_type = "rental_based"
    elif "灯光架" in text: price = 2000; item_type = "rental_based"
    elif "信号放大器" in text: price = 300; item_type = "rental_based"
    elif "灯光控台" in text: price = 2000; item_type = "rental_based"
    elif "灯光师" in text: price = 1200; item_type = "labor"
    
    # 4. Performers / Other
    elif "提琴" in text: price = 1500; item_type = "labor"
    elif "手绘师" in text: price = 2500; item_type = "labor"
    elif "乐队" in text: price = 6000; item_type = "labor"
    elif "主持人" in text: price = 3500; item_type = "labor"
    elif "儿童演员" in text: price = 600; item_type = "labor"
    elif "高空球" in text: price = 4000; item_type = "rental_based"
    elif "吊车" in text: price = 5000; item_type = "rental_based"
    elif "高空杂技" in text: price = 2500; item_type = "labor"
    elif "走秀模特" in text: price = 1800; item_type = "labor"
    elif "化妆师" in text: price = 1200; item_type = "labor"
    elif "节目背景" in text: price = 1500; item_type = "purchase_based"
    elif "摄影师" in text: price = 1500; item_type = "labor"
    elif "图片直播" in text: price = 2500; item_type = "labor"
    elif "摄像" in text: price = 1800; item_type = "labor"
    elif "航拍" in text: price = 2000; item_type = "labor"
    elif "高清机位" in text: price = 2000; item_type = "labor"
    elif "剪辑花絮" in text: price = 1500; item_type = "labor"
    elif "礼服" in text: price = 5000; item_type = "rental_based"
    elif "非遗簪花" in text: price = 8000; item_type = "labor"
    elif "非遗鱼灯" in text: price = 150; item_type = "purchase_based"
    elif "运输费" in text: price = 1000; item_type = "labor"
    elif "人工费" in text: price = 500; item_type = "labor"

    # Adjust Price based on Form
    if item_type == "rental_based":
        if "制作" in form or "购买" in form:
            price = price * 5 # Buying is much more expensive than renting
        elif "租赁" in form:
            price = price * 1 # Base is rental
            
    elif item_type == "purchase_based":
        if "租赁" in form:
            price = price * 0.3 # Renting a "make" item? (Rare, but maybe reusing frames)
        elif "制作" in form or "购买" in form:
            price = price * 1 # Base is purchase
            
    return price

# Load the ORIGINAL file to ensure clean state for column insertion
# Or load the UPDATED one if we want to preserve previous fixes?
# User said "Original Unit Price survey might be inaccurate", so re-calculating is good.
# User previously asked to fix images in K column (which was J in original?). 
# Actually, K column in original? 
# Original: A,B,C,D,E,F,G,H,I,J (10 cols). K is 11. Maybe images were in K?
# Let's use the file we just fixed: _updated.xlsx
file_path = r'g:\learning\agent-skills\xmls\（标段1：活动）营销中心样板房开放报价清单_updated.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

print(f"Original images count: {len(sheet._images)}")

# Insert Column I (Index 9)
print("Inserting column at index 9...")
sheet.insert_cols(9)

# Set Header for new I column (Row 2, based on observation)
sheet.cell(row=2, column=9).value = "天数（无需计算填1）"

# Iterate and Update
# New Column Mapping:
# F (6): Qty
# G (7): Unit
# H (8): Form
# I (9): Days (New)
# J (10): Unit Price (Was I)
# K (11): Total Price (Was J)

for row in range(3, sheet.max_row + 1):
    # Check if it's a data row (has Qty)
    qty_cell = sheet.cell(row=row, column=6)
    qty = qty_cell.value
    
    if qty is None:
        continue
        
    # Set Days = 1
    sheet.cell(row=row, column=9).value = 1
    
    # Get Data for Pricing
    item_name = sheet.cell(row=row, column=3).value
    material = sheet.cell(row=row, column=4).value
    spec = sheet.cell(row=row, column=5).value
    unit = sheet.cell(row=row, column=7).value
    form = sheet.cell(row=row, column=8).value
    
    if item_name is None: continue

    # Recalculate Unit Price
    unit_price = estimate_price(item_name, material, form, unit, spec)
    
    # Write Unit Price to J (10)
    sheet.cell(row=row, column=10).value = unit_price
    
    # Write Total Price Formula to K (11)
    # Formula: F * J * I
    sheet.cell(row=row, column=11).value = f"=F{row}*J{row}*I{row}"

# Save
new_file_path = file_path # Overwrite the updated file or create new? 
# Let's overwrite the _updated.xlsx to avoid chain of files, or make _v2.
new_file_path = r'g:\learning\agent-skills\xmls\（标段1：活动）营销中心样板房开放报价清单_v2.xlsx'
wb.save(new_file_path)
print(f"Processing complete. Saved to {new_file_path}")

# Verify images
wb_new = openpyxl.load_workbook(new_file_path)
print(f"Saved file images count: {len(wb_new.active._images)}")
